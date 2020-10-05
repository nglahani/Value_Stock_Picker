# -*- coding: utf-8 -*-33
"""
Created on Thu Nov 14 15:15:10 2019
Collects Stock Data on All Available Entities in SimFin DataBase

@author: Niko Lahanis and George Haire
"""

import requests
import bs4 as bs
import pickle 
import math
import openpyxl

def pietroski(c):
    l=[]
    for item in c:
        score = item[11].get('value')
        if score is None:
            score=0
        score=float(score)
        l.append(score) 
    return l
        
def divPerShare(c):
    l=[]
    for item in c: 
        score = item[7].get('value')
        if score is None:
            score=0
        score=float(score)
        if score > 0:
            score = 5
        else:
            score = 0
        l.append(score) 
    return l

def PriceBook(c):
    l=[]
    for item in c:
        score = item[9].get('value')
        if score is None:
            score=0
        score=float(score)
        if score > 5:
            score = 1
        elif 3 < score <= 5:
            score = 2
        elif 2 < score <= 3:
            score =3
        elif 1 < score <= 2:
            score =4
        elif 0 < score <= 1:
            score = 5
        else:
            score = 0
        l.append(score)  
    return l

def CurrentRatio(c):
    l=[]
    for item in c:
        score = item[4].get('value')
        if score is None:
            score=0
        score=float(score)
        if 0 < score <= 1:
            score = 1
        elif score > 3:
            score = 1
        elif 2 < score <= 3:
            score =3
        elif 1.5 < score <= 2:
            score =4
        elif 1 < score <= 1.5:
            score = 5
        else:
            score = 0
        l.append(score)
    return l

def EPS(c):
    l=[]
    for item in c:
        score = item[5].get('value')
        if score is None:
            score=0
        score=float(score)
        if score < 0:
            score = 0
        l.append(score)
    return l


def PricetoEarnings(c):
    l=[]
    for item in c:
        score = item[8].get('value')
        if score is None:
            score=0
        score=float(score)
        if score > 30:
            score = 1
        elif 21 < score <= 30:
            score = 2
        elif 16 < score <= 20:
            score =3
        elif 10 < score <= 15:
            score =4
        elif 1 < score <= 10:
            score = 5
        else:
            score = 0
        l.append(score)
    return l

def BookperShare(c):
    l=[]
    for item in c:
        score = item[6].get('value')
        if score is None:
            score=0
        score=float(score)
        if score < 0:
            score = 0
        l.append(score)
    return l

def save_sp500_tickers():
    resp = requests.get('http://en.wikipedia.org/wiki/List_of_S%26P_500_companies')
    soup = bs.BeautifulSoup(resp.text, 'lxml')
    table = soup.find('table', {'class': 'wikitable sortable'})
    tickers = []
    for row in table.findAll('tr')[1:]:
        ticker = row.findAll('td')[0].text
        tickers.append(ticker)
    
    with open("sp500tickers.pickle","wb") as f:
        pickle.dump(tickers,f)
        
    return tickers

sp500=save_sp500_tickers()
sp500=str(sp500)

api_key="mwQjjLorlcNiWr3Oi9O3G0nQDDz6s4Mc"

uRL = "https://simfin.com/api/v1/info/all-entities?api-key={}".format(api_key)
response = requests.get(uRL)
dataID = response.json()   

a=[]
b=[]
for item in dataID:
    ticker=item.get('ticker')
    if ticker is None:
        item['ticker']=''
        ticker=item.get('ticker')
    ticker=ticker+'\\n'
    if ticker in sp500 and ticker != '\\n':
        simId=item.get('simId')
        a.append(simId)
        b.append(ticker)


"""
Stock Indicator List:
    0-73 - Sector Classification
    0-71 - Ticker
    0-31 - Last Close Price
    4-11 - Market Capitalisation
    4-3 - Current Ratio
    4-13 - Earnings per Share, Diluted
    4-18 - Book Value per Share
    4-29 - Dividends per Share
    4-14 - Price to Earnings Ratio
    4-16 - Price to Book Value
    4-21 - EV/EBITDA
    4-30 - Pietroski F-Score
"""
indicators=["0-73","0-71","0-31","4-11","4-3","4-13","4-18","4-29","4-14","4-16","4-21","4-30"]
indicatorStr=','.join(indicators)


"""
In this next section, pertinent stock metrics for each stock in SimFin's database will be 
pulled using another API request. Each stock will be given a unique index found in list b.
"""

"""
Test Run for one set of Stock Data from Google":
    
dataURL="https://simfin.com/api/v1/companies/id/{}/ratios?api-key={}&indicators={}".format(18,api_key,indicatorStr)
dataResponse=requests.get(dataURL)
singleStockData = dataResponse.json()
print(singleStockData)
"""

c=[]
for item in a:
    dataURL="https://simfin.com/api/v1/companies/id/{}/ratios?api-key={}&indicators={}".format(item,api_key,indicatorStr)
    dataResponse=requests.get(dataURL)
    singleStockData = dataResponse.json()
    c.append(singleStockData)

"""
We now have all the data stored in a json file format with variable c
"""
mOS = [0]*len(c)


pietroskiList = pietroski(c)
divList = divPerShare(c)
PBList = PriceBook(c)
currentList = CurrentRatio(c)
eList = EPS(c)
PEList = PricetoEarnings(c)
bShareList = BookperShare(c)

"""
All the Basic Metrics are now in a list of values
"""

grahamNumber = [math.sqrt(22.5*a * b) for a, b in zip(eList, bShareList)]

"""
Graham Number is Calculated
"""

excel={"Ticker":["Pietroski","Dividends Per Share", "Price to Book", "Current Ratio", "Price to Earnings", "Graham Number","Total"]}
i=0

for item in c:
    finalList=[pietroskiList[i],divList[i],PBList[i],currentList[i],PEList[i],grahamNumber[i]]
    finalList.append(sum(finalList[:5]))
    excel.update({item[1].get('value'):finalList})
    i=i+1
   
        
"""
Stored into a Dictionary to Export into Excel
"""

# Create the workbook and sheet for Excel
workbook = openpyxl.Workbook()
sheet = workbook.active

row = 1
for key,values in excel.items():
    # Put the key in the first column for each key in the dictionary
    sheet.cell(row=row, column=1, value=key)
    column = 2
    for element in values:
        # Put the element in each adjacent column for each element in the tuple
        sheet.cell(row=row, column=column, value=element)
        column += 1
    row += 1

workbook.save(filename="StockPicker.xlsx")

"""
Data Imported to an Excel File
"""
 
